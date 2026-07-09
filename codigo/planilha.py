"""Leitura/escrita do workbook de mercado (.xlsx) — saída estilizada e por colunas.

Cada campo em sua coluna (sem ambiguidade de separador como no CSV aberto no Excel).
Cabeçalho formatado e painel congelado. Escrita atômica (tmp + os.replace) com retry
para o lock intermitente do OneDrive no Windows.
"""
from __future__ import annotations

import os
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="305496")


def ler_rows(path, aba: str) -> list[dict]:
    """Lê uma aba do workbook -> lista de dicts (cabeçalho como chaves). [] se não houver."""
    p = Path(path)
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            return []
        it = wb[aba].iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return []
        cols = [c for c in header if c is not None]
        return [
            {col: r[i] for i, col in enumerate(cols)}
            for r in it
            if r is not None and r[0] is not None
        ]
    finally:
        wb.close()


def _estilizar(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"


def escrever_mercado(path, abas) -> None:
    """Escreve o workbook. Atômico + estilizado.

    `abas` = lista de tuplas `(nome, colunas, rows)` ou
    `(nome, colunas, rows, formatos)`, onde `formatos` é um dict
    `{nome_coluna: number_format}` (ex.: 8 casas decimais "0.00000000").
    Valores que começam com "=" são gravados como fórmula do Excel.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for aba in abas:
        nome, colunas, rows = aba[0], aba[1], aba[2]
        formatos = aba[3] if len(aba) > 3 else {}
        ws = wb.create_sheet(nome)
        ws.append(list(colunas))
        for r in rows:
            ws.append([r.get(c) for c in colunas])
        _estilizar(ws, len(colunas))
        for col_nome, fmt in formatos.items():
            if col_nome in colunas:
                ci = colunas.index(col_nome) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=ci).number_format = fmt

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_name(p.name + ".tmp")
    wb.save(tmp)
    for tentativa in range(6):
        try:
            os.replace(tmp, p)
            return
        except PermissionError as erro:
            if tentativa == 5:
                raise PermissionError(
                    f"não consegui gravar {p.name}: acesso negado — o arquivo "
                    f"provavelmente está ABERTO no Excel. Feche-o e rode de novo. "
                    f"(conteúdo novo deixado em {tmp.name})"
                ) from erro
            time.sleep(0.5 * (tentativa + 1))
